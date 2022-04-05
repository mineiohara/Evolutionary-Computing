import random
import math
from typing import List
import openpyxl


class Gene(object):
    def __init__(self, realValuedGene: List[float]) -> None:
        self.realValuedGene = realValuedGene
        self.eval = None


class Generation(object):
    # Initialize
    def __init__(self, popSize: int, dimention: int, pc: float, a: float, pm: float, b: int, maxGeneration: int, searchSpace: List[float], decimalPoints: int) -> None:
        self.popSize = popSize
        self.dimention = dimention
        self.pc = pc
        self.a = a
        self.pm = pm
        self.b = b
        self.currentGeneration = None
        self.maxGeneration = maxGeneration
        self.searchSpace = searchSpace
        self.decimalPoints = decimalPoints
        self.population = [ Gene([round(random.random()*(searchSpace[1] - searchSpace[0])+ searchSpace[0], decimalPoints) for _ in range(dimention)]) for _ in range(popSize) ]
        self.minEval = [0, 100000000000000]

    # Evaluation function 1
    def schewefel(self) -> None:
        for gene in self.population:
            sum = 0
            product = 1
            for i in range(self.dimention):
                sum += abs(gene.realValuedGene[i])
                product *= abs(gene.realValuedGene[i])
            gene.eval = sum + product
            if abs(gene.eval) < abs(self.minEval[1]):
                self.minEval[0] = self.currentGeneration
                self.minEval[1] = gene.eval

    # Evaluation function 2
    def griewank(self) -> None:
        for gene in self.population:
            sum = 0
            product = 1
            for i in range(self.dimention):
                sum += gene.realValuedGene[i]**2
                product *= math.cos(gene.realValuedGene[i]/math.sqrt(i+1))
            gene.eval = 1/4000*sum - product + 1
            if abs(gene.eval) < abs(self.minEval[1]):
                self.minEval[0] = self.currentGeneration
                self.minEval[1] = gene.eval

    # Non-uniform mutation
    def mutation(self) -> None:
        for gene in self.population:
            if random.random() < self.pm:
                if random.randint(0,1):
                    for i in range(self.dimention):
                        gene.realValuedGene[i] -= (gene.realValuedGene[i] - self.searchSpace[0]) * (1 - (random.random() ** ((1 - self.currentGeneration/self.maxGeneration) ** self.b)))
                    gene.realValuedGene[i] = round(gene.realValuedGene[i], self.decimalPoints)

                else:
                    for i in range(self.dimention):
                        gene.realValuedGene[i] += (self.searchSpace[1] - gene.realValuedGene[i]) * (1 - (random.random() ** ((1 - self.currentGeneration/self.maxGeneration) ** self.b)))
                    gene.realValuedGene[i] = round(gene.realValuedGene[i], self.decimalPoints)

    # Whole arithmetic crossover
    def crossover(self) -> None:
        isCrossover = []
        for _ in range(self.popSize):
            if random.random() <= self.pc: isCrossover.append(True)
            else: isCrossover.append(False)

        for i in range(self.popSize):
            if isCrossover[i]:
                for j in range(i+1, self.popSize):
                    if isCrossover[j]:
                        isCrossover[j] = False
                        child = []
                        for k in range(self.dimention):
                            child.append(round(self.population[i].realValuedGene[k] * self.a + self.a * self.population[j].realValuedGene[k], self.decimalPoints))
                        
                        self.population.append(Gene(child))
                        self.population.append(Gene(child))
                        del child

    # Tournament selection
    def selection(self) -> None:
        oldGen = self.population
        self.population = []
        for _ in range(self.popSize):
            rand1 = random.randint(0, len(oldGen) - 1)
            rand2 = random.randint(0, len(oldGen) - 1)

            if abs(oldGen[rand1].eval) < abs(oldGen[rand2].eval):
                self.population.append(oldGen[rand1])
            else: 
                self.population.append(oldGen[rand2])
        
        del oldGen



if __name__=='__main__':
    #Set parameters
    popSize = 20
    dimention = 30
    searchSpace = [-10, 10]
    pc = 0.25
    a = 0.5
    pm = 0.4
    maxGeneration = 200000
    b = 5
    decimalPoints = 9
    acceptance = 0.01

    # For evaluation function 1
    for m in range(100):
        g = Generation(popSize, dimention, pc, a, pm, b, maxGeneration, searchSpace, decimalPoints)
        for i in range(g.maxGeneration):
            g.currentGeneration = i+1
            g.crossover()
            g.mutation()
            g.schewefel()
            g.selection()
        
        print("Optimum value: {0} at generation {1}".format(g.minEval[1], g.minEval[0]))
        wb = openpyxl.load_workbook('./data.xlsx')
        sheet = wb['Sheet1']
        sheet.cell(row=m+1, column=1).value = m+1
        if g.minvalue[1] != 0: sheet.cell(row=m+1, column=2).number_format = '0.000000000'
        sheet.cell(row=m+1, column=2).value = g.minEval[1]*100
        sheet.cell(row=m+1, column=3).value = g.minEval[0]
        wb.save('./data.xlsx')
        wb.close()
        del g

    
    # For evaluation function 2
    pm = 1
    searchSpace = [-600, 600]

    for m in range(100):
        g = Generation(popSize, dimention, pc, a, pm, b, maxGeneration, searchSpace, decimalPoints)
        for i in range(g.maxGeneration):
            g.currentGeneration = i+1
            g.crossover()
            g.mutation()
            g.griewank()
            g.selection()
            if g.minEval[1] == 0: break
        
        print("Good!", end=" ") if g.minEval[1] < acceptance else print("Bad!", end=" ")
        print("Optimum value: {0} at generation {1}".format(g.minEval[1], g.minEval[0]))
        wb = openpyxl.load_workbook('./data2.xlsx')
        sheet = wb['Sheet1']
        sheet.cell(row=m+1, column=1).value = m+1
        if g.minEval[1] != 0: sheet.cell(row=m+1, column=2).number_format = '0.000000000'
        sheet.cell(row=m+1, column=2).value = g.minEval[1]
        sheet.cell(row=m+1, column=3).value = g.minEval[0]
        wb.save('./data2.xlsx')
        wb.close()
        del g