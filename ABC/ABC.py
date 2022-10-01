from cmath import inf
import random
import numpy as np
from matplotlib import pyplot as plt
from typing import List
import matplotlib.cm as cm
import time
import openpyxl


class Benchmark(object):
    def __init__(self, fSet: 'FlowerSet', funcName: str) -> int:
        self.fSet = fSet
        self.funcName = funcName
        self.funcList = {'schwefel':[500,-500,0.01], 'ackley':[32,-32,0.01], 'sphere':[100,-100,0.01], 'quadricNoise':[100,-100,100], 'rosenbrock':[10,-10,100], 'griewank':[600,-600,0.01]}
        self.fSet.COEF_MAX = self.funcList[self.funcName][0]
        self.fSet.COEF_MIN = self.funcList[self.funcName][1]
        if DIMENSION == 1 or DIMENSION == 2:
            self.fig, ax = plt.subplots()
            plt.ion()
            if DIMENSION == 1:
                x = np.linspace(fSet.COEF_MIN, fSet.COEF_MAX, int((fSet.COEF_MAX-fSet.COEF_MIN)*10))
                plt.xlabel("x",fontsize=18)
                plt.ylabel("f(x)",fontsize=18)
                y = self.evaluate(x)
                ax.plot(x, y, linewidth=0.75)
            elif DIMENSION == 2:
                x1 = np.linspace(fSet.COEF_MIN, fSet.COEF_MAX, fSet.COEF_MAX-fSet.COEF_MIN)
                x2 = np.linspace(fSet.COEF_MIN, fSet.COEF_MAX, fSet.COEF_MAX-fSet.COEF_MIN)
                dy = []
                for i in x1:
                    temp = []
                    for j in x2:
                        temp.append(self.evaluate([i,j]))
                    dy.append(temp)
                heatmap = ax.pcolor(dy, cmap=cm.jet)
                plt.colorbar(heatmap)
                ax.set_xticklabels(np.arange(fSet.COEF_MIN, fSet.COEF_MAX, fSet.COEF_MAX-fSet.COEF_MIN))
                ax.set_yticklabels(np.arange(fSet.COEF_MIN, fSet.COEF_MAX, fSet.COEF_MAX-fSet.COEF_MIN))
            self.sc = ax.scatter([], [], color="yellow", s=20, edgecolors="black")
            self.bsc = ax.scatter([], [], color="red", s=20, edgecolors="black")

    def plot(self):
        if DIMENSION == 1: self.plot1D()
        elif DIMENSION == 2: self.plot2D()

    def plot1D(self):
        x, f = [], []
        for i in range(EBEE_NUM):
            x.append(self.fSet.flower[i].pos[0])
            f.append(self.fSet.flower[i].value)

        b1 = [self.fSet.bestPos[0]]
        b2 = [self.fSet.bestValue]

        self.sc.set_offsets(np.c_[x,f])
        self.bsc.set_offsets(np.c_[b1,b2])
        self.fig.canvas.draw_idle()
        plt.pause(WAIT_TIME)

    def plot2D(self):
        x1, x2 = [], []
        for i in range(EBEE_NUM):
            x1.append(self.fSet.flower[i].pos[0] + self.fSet.COEF_MAX)
            x2.append(self.fSet.flower[i].pos[1] + self.fSet.COEF_MAX)

        b1 = self.fSet.bestPos[0] + self.fSet.COEF_MAX
        b2 = self.fSet.bestPos[1] + self.fSet.COEF_MAX

        self.sc.set_offsets(np.c_[x1,x2])
        self.bsc.set_offsets(np.c_[b1,b2])
        self.fig.canvas.draw_idle()
        plt.pause(WAIT_TIME)
        

    def evaluate(self, pos: List[float]) -> float:
        func = getattr(self, self.funcName)
        return func(pos)

    def schwefel(self, pos: List[float]) -> float:
        if type(pos) == np.ndarray:
            return 418.9829 * DIMENSION - pos * np.sin(np.sqrt(np.abs(pos)))
        sum = 0
        for i in range(DIMENSION):
            sum += pos[i] * np.sin(np.sqrt(np.abs(pos[i])))
        return 418.9829 * DIMENSION - sum

    def ackley(self, pos: List[float]) -> float:
        if type(pos) == np.ndarray:
            return -20 * np.exp(-0.2*np.sqrt(1/DIMENSION*(pos**2))) - np.exp(1/DIMENSION*np.cos(2*np.pi*pos)) + 20 + np.e
        sum1, sum2 = 0, 0
        for i in range(DIMENSION):
            sum1 += pos[i] ** 2
            sum2 += np.cos(2 * np.pi * pos[i])
        return -20 * np.exp(-0.2*np.sqrt(1/DIMENSION*sum1)) - np.exp(1/DIMENSION*sum2) + 20 + np.e
    
    def sphere(self, pos: List[float]):
        if type(pos) == np.ndarray:
            return pos**2
        sum = 0
        for i in range(DIMENSION):
            sum += pos[i]**2
        return sum

    def quadricNoise(self, pos: List[float]):
        if type(pos) == np.ndarray:
            return pos**2
        sum2 = 0
        for i in range(DIMENSION):
            sum1 = 0
            for j in range(0, i+1):
                sum1 += pos[j]
            sum2 += sum1**2
        return sum2

    def rosenbrock(self, pos: List[float]):
        if type(pos) == np.ndarray:
            return 100 * ((pos**2)**2) + (pos - 1)**2
        sum = 0
        for i in range(DIMENSION-1):
            sum += 100 * ((pos[i+1] - pos[i]**2)**2) + (pos[i]-1)**2
        return sum

    def griewank(self, pos: List[float]):
        if type(pos) == np.ndarray:
            return 1/4000 * (pos**2) - np.cos(pos) + 1
        sum = 0
        product = 1
        for i in range(DIMENSION):
            sum += pos[i]**2
            product *= np.cos(pos[i]/(i+1))
        return 1/4000 * sum - product + 1
    



class Flower(object):
    def __init__(self, fSet: 'FlowerSet') -> None:
        self.fSet = fSet
        self.value = None
        self.pos = [(fSet.COEF_MIN + (fSet.COEF_MAX - fSet.COEF_MIN) * random.random()) for _ in range(DIMENSION)]
        self.visitNum = 0
        self.evaluate()

    def change(self, base: int):
        for i in range(DIMENSION):
            self.pos[i] = self.fSet.flower[base].pos[i]

        i = random.randint(0, DIMENSION-1)
        j = (base + (random.randint(0, EBEE_NUM-2) + 1)) % EBEE_NUM
        self.pos[i] += (random.random() * 2 - 1) * (self.pos[i] - self.fSet.flower[j].pos[i])
        if self.pos[i] > self.fSet.COEF_MAX:
            self.pos[i] = self.fSet.COEF_MAX
        elif self.pos[i] < self.fSet.COEF_MIN:
            self.pos[i] = self.fSet.COEF_MIN
        self.visitNum = 0
        self.evaluate()

    def renew(self):
        for i in range(DIMENSION):
            self.pos[i] = self.fSet.COEF_MIN + (self.fSet.COEF_MAX - self.fSet.COEF_MIN) * random.random()
        self.visitNum = 0
        self.evaluate()

    def evaluate(self):
        self.value = self.fSet.benchmark.evaluate(self.pos)

class FlowerSet(object):
    def __init__(self, funcName: str) -> None:
        self.COEF_MAX = None
        self.COEF_MIN = None
        self.benchmark = Benchmark(self, funcName)
        self.flower = [Flower(self) for _ in range(EBEE_NUM)]
        self.best = 0
        self.bestPos = []
        self.newFlower = Flower(self)
        self.trValue = [0 for _ in range(EBEE_NUM)]
        self.bestValue = float(inf)

        newBest = False
        for i in range(EBEE_NUM):
            if self.bestValue > self.flower[i].value:
                self.best = i
                newBest = True
        if newBest:
            self.bestPos = self.flower[self.best].pos[:]
            self.bestValue = self.flower[self.best].value
            

    def employedBeePhase(self):
        for i in range(EBEE_NUM):
            self.newFlower.change(i)
            if self.flower[i].value > self.newFlower.value:
                self.newFlower, self.flower[i] = self.flower[i], self.newFlower
            self.flower[i].visitNum += 1

    def onlookerBeePhase(self):
        for _ in range(OBEE_NUM):
            max = -float(inf)
            min = float(inf)
            for i in range(EBEE_NUM):
                if max < self.flower[i].value:
                    max = self.flower[i].value
                if min > self.flower[i].value:
                    min = self.flower[i].value
        
            denom = 0.0
            for i in range(EBEE_NUM):
                self.trValue[i] = (max - self.flower[i].value) / (max - min)
                denom += self.trValue[i]
            
            r = random.random()
            for i in range(EBEE_NUM - 1):
                prob = self.trValue[i] / denom
                if r <= prob:
                    temp = i
                    break
                temp = i
                r -= prob
            
            self.newFlower.change(temp)
            if self.flower[temp].value > self.newFlower.value:
                self.newFlower, self.flower[temp] = self.flower[temp], self.newFlower

            self.flower[temp].visitNum += 1


    def scoutBeePhase(self):
        for i in range(EBEE_NUM):
            if VISIT_MAX <= self.flower[i].visitNum:
                self.flower[i].renew()
    
    def saveBestPos(self):
        newBest = False
        for i in range(EBEE_NUM):
            if self.bestValue > self.flower[i].value:
                self.best = i
                newBest = True
        if newBest:
            self.bestPos = self.flower[self.best].pos[:]
            self.bestValue = self.flower[self.best].value

    def printResult(self, funcName: str):
        print('{1}: f(x) = {0}'.format(self.bestValue, funcName))


def main(funcName: str) -> None:
    fSet = FlowerSet(funcName)
    # start = time.time()
    for i in range(1, REPEAT_NUM+1):
        # plt.title("Step {0}".format(i),fontsize=25)
        start = time.time()
        fSet.employedBeePhase()
        fSet.onlookerBeePhase()
        fSet.scoutBeePhase()
        fSet.saveBestPos()
        # fSet.benchmark.plot()
        # if i == 1:
        #     input()
        end = time.time()
        print('Elapsed Time: {0}'.format((end - start)*1000))
        # if fSet.bestValue < fSet.benchmark.funcList[funcName][2]:
        print('{2} {0} 回目：最良評価値 {1}'.format(i, fSet.bestValue, funcName))
            # break

        # wb = openpyxl.load_workbook('./data2.xlsx')
        # sheet = wb['Sheet1']
        # sheet.cell(row=i, column=1).value = i
        # if fSet.bestValue != 0: sheet.cell(row=i, column=2).number_format = '0.000000000'
        # sheet.cell(row=i, column=2).value = fSet.bestValue
        # wb.save('./data2.xlsx')
        # wb.close()
    
    # print("Elapsed time: {0}".format(time.time() - start))
    
    fSet.printResult(funcName)
    del(fSet)

if __name__ == '__main__':
    REPEAT_NUM = 200000
    DIMENSION = 30
    MAGNIFICATION = 0.6
    EBEE_NUM = int(31 * DIMENSION * MAGNIFICATION) #200, 30, 150
    OBEE_NUM = int(4 * DIMENSION * MAGNIFICATION)
    VISIT_MAX = int(92 * DIMENSION * MAGNIFICATION)
    WAIT_TIME = 0.00001

    funcList = ['schwefel','ackley','sphere','quadricNoise','rosenbrock','griewank']

    for _ in range(10): main('griewank')
    # for i in range(100):
    #     print(i+1)
    #     for func in funcList:
    #         main(func)